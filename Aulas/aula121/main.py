"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""
import openpyxl
from random import uniform
"""
pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Planilha1']

#adicionando valores na planilha
for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')
"""

"""
#alterando o valor de b3
planilha1['b3'].value = 2200
#tem que salvar o arquivo
"""


"""
#ler conteudo da planilha
#print(planilha1['b'].value)

for coluna in planilha1:
    print(coluna[0].value, coluna[1].value, coluna[2].value)
"""
planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'Kelvin {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 2).value = f'Otávio {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'maria {linha} {round(uniform(10, 100), 2)}'

planilha.save('nova_planilha1.xlsx')
